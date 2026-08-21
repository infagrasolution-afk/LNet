import uuid
import csv
import io
from datetime import datetime
from typing import List, Optional
from fastapi import FastAPI, HTTPException, Depends, Body, Response
from fastapi.responses import HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

import store
import mailer

app = FastAPI(title="LNet Backend API")

# Enable CORS for React frontend (Vite dev server)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Models
class LoginRequest(BaseModel):
    username: str
    password: str

class UserCreateRequest(BaseModel):
    name: str
    cedula: str
    username: Optional[str] = None
    password: Optional[str] = None
    role: str = "user"

class UserStatusRequest(BaseModel):
    status: str  # "activo" or "bloqueado"

class SettingsRequest(BaseModel):
    gmail_user: str
    gmail_app_password: str
    default_recipients: Optional[str] = ""

class TestEmailRequest(BaseModel):
    recipient: str

class ActivityItem(BaseModel):
    id: str
    name: str
    description: Optional[str] = ""
    detail: Optional[str] = ""
    checked: bool = False
    unid_mts: Optional[str] = ""

class RecordCreateRequest(BaseModel):
    solicitud_num: str
    client_name: str
    activities: List[ActivityItem]
    observations: Optional[str] = ""
    created_by: str
    send_email: bool = False
    recipient_email: Optional[str] = None

@app.on_event("startup")
def startup_event():
    store.ensure_data_dir()

# Auth Endpoint
@app.post("/api/auth/login")
def login(req: LoginRequest):
    users = store.load_users()
    username_clean = req.username.strip().lower()
    
    user = next((u for u in users if u["username"].lower() == username_clean), None)
    if not user:
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos.")
        
    if user["password"] != req.password:
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos.")
        
    if user.get("status") == "bloqueado":
        raise HTTPException(
            status_code=403, 
            detail="Usuario bloqueado. Contacte al administrador del sistema."
        )

    return {
        "message": "Inicio de sesión exitoso",
        "user": {
            "id": user["id"],
            "name": user["name"],
            "username": user["username"],
            "cedula": user["cedula"],
            "role": user["role"],
            "status": user["status"]
        }
    }

# Admin User Management Endpoints
@app.get("/api/users")
def get_users():
    users = store.load_users()
    # Don't send passwords in list
    clean_users = [
        {
            "id": u["id"],
            "name": u["name"],
            "username": u["username"],
            "cedula": u["cedula"],
            "role": u["role"],
            "status": u.get("status", "activo"),
            "created_at": u.get("created_at", "")
        }
        for u in users
    ]
    return clean_users

@app.post("/api/users")
def create_user(req: UserCreateRequest):
    users = store.load_users()
    
    # Format username if not provided: initial + surname
    if not req.username:
        name_parts = req.name.strip().split()
        if len(name_parts) >= 2:
            initial = name_parts[0][0].lower()
            lastname = name_parts[-1].lower()
            gen_username = f"{initial}{lastname}"
        else:
            gen_username = req.name.strip().lower()
    else:
        gen_username = req.username.strip().lower()

    if any(u["username"].lower() == gen_username for u in users):
        raise HTTPException(status_code=400, detail=f"El nombre de usuario '{gen_username}' ya existe.")

    # Password defaults to cedula if empty
    gen_password = req.password if req.password else req.cedula.strip()

    new_user = {
        "id": str(uuid.uuid4()),
        "name": req.name.strip(),
        "cedula": req.cedula.strip(),
        "username": gen_username,
        "password": gen_password,
        "role": req.role,
        "status": "activo",
        "created_at": datetime.now().isoformat()
    }
    
    users.append(new_user)
    store.save_users(users)
    return {"message": "Usuario creado exitosamente", "user": new_user}

@app.put("/api/users/{username}/status")
def update_user_status(username: str, req: UserStatusRequest):
    users = store.load_users()
    user = next((u for u in users if u["username"].lower() == username.lower()), None)
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")
        
    if user["role"] == "admin" and req.status == "bloqueado":
        raise HTTPException(status_code=400, detail="No se puede bloquear al usuario administrador principal.")

    user["status"] = req.status
    store.save_users(users)
    return {"message": f"Estado de usuario actualizado a {req.status}", "status": req.status}

@app.post("/api/users/{username}/reset-password")
def reset_password(username: str):
    users = store.load_users()
    user = next((u for u in users if u["username"].lower() == username.lower()), None)
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")

    user["password"] = user["cedula"]
    store.save_users(users)
    return {"message": f"Contraseña de {username} restablecida a su número de cédula ({user['cedula']})."}

@app.delete("/api/users/{username}")
def delete_user(username: str):
    users = store.load_users()
    user = next((u for u in users if u["username"].lower() == username.lower()), None)
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")
        
    if user["role"] == "admin":
        raise HTTPException(status_code=400, detail="No se puede eliminar al usuario administrador principal.")

    filtered_users = [u for u in users if u["username"].lower() != username.lower()]
    store.save_users(filtered_users)
    return {"message": f"Usuario {username} eliminado exitosamente."}

# Settings Endpoints
@app.get("/api/settings")
def get_settings():
    return store.load_settings()

@app.post("/api/settings")
def save_settings(req: SettingsRequest):
    settings = {
        "gmail_user": req.gmail_user.strip(),
        "gmail_app_password": req.gmail_app_password.strip(),
        "default_recipients": req.default_recipients.strip() if req.default_recipients else ""
    }
    store.save_settings(settings)
    return {"message": "Configuración guardada exitosamente", "settings": settings}

@app.post("/api/test-email")
def test_email(req: TestEmailRequest):
    settings = store.load_settings()
    sender = settings.get("gmail_user")
    app_pw = settings.get("gmail_app_password")

    if not sender or not app_pw:
        raise HTTPException(
            status_code=400, 
            detail="Debe configurar primero el correo y la contraseña de aplicación de Gmail en el panel admin."
        )

    dummy_record = {
        "solicitud_num": "999999",
        "client_name": "PRUEBA DE CONEXIÓN GMAIL",
        "created_by": "Sistema Admin",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "activities": [
            {"description": "Prueba de conexión con Gmail SMTP", "checked": True, "unid_mts": "1 OK"}
        ],
        "observations": "Correo de prueba enviado desde el Panel Administrativo de LNet."
    }

    try:
        mailer.send_gmail_email(
            sender_email=sender,
            app_password=app_pw,
            recipient_emails=req.recipient,
            subject="[LNet] Prueba de Conexión de Correo Gmail Exitosa",
            record_data=dummy_record
        )
        return {"message": "Correo de prueba enviado exitosamente a " + req.recipient}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar correo por Gmail: {str(e)}")

# Records Form Endpoints
@app.get("/api/records")
def get_records(username: Optional[str] = None):
    records = store.load_records()
    if username:
        return [r for r in records if r.get("created_by", "").lower() == username.lower()]
    return records

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

@app.get("/api/records/export/excel")
def export_records_excel(username: Optional[str] = None, record_id: Optional[str] = None):
    records = store.load_records()
    if record_id:
        records = [r for r in records if r.get("id") == record_id]
    elif username:
        records = [r for r in records if r.get("created_by", "").lower() == username.lower()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte LNet"

    # Header style
    header_fill = PatternFill(start_color="01579B", end_color="01579B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    headers = [
        "Nro. Solicitud",
        "Cliente / Razón Social",
        "Registrado Por",
        "Fecha / Hora",
        "Materiales / Actividades Ejecutadas",
        "Observaciones"
    ]
    ws.append(headers)

    # Style header row
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    ws.row_dimensions[1].height = 26

    # Append data rows
    for row_idx, r in enumerate(records, start=2):
        executed_acts = [
            f"• {a.get('description', a.get('name'))}" + (f" ({a.get('detail')})" if a.get('detail') else "") + f" [{a.get('unid_mts', '1')} unid/mts]"
            for a in r.get("activities", []) if a.get("checked")
        ]
        acts_str = "\n".join(executed_acts) if executed_acts else "Sin materiales marcados"

        row_data = [
            r.get("solicitud_num", ""),
            r.get("client_name", ""),
            r.get("created_by", ""),
            r.get("created_at", ""),
            acts_str,
            r.get("observations", "")
        ]
        ws.append(row_data)

        # Style cells
        ws.cell(row=row_idx, column=1).alignment = center_align
        ws.cell(row=row_idx, column=2).alignment = left_align
        ws.cell(row=row_idx, column=3).alignment = center_align
        ws.cell(row=row_idx, column=4).alignment = center_align
        ws.cell(row=row_idx, column=5).alignment = left_align
        ws.cell(row=row_idx, column=6).alignment = left_align

        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_num).border = thin_border

        ws.row_dimensions[row_idx].height = max(30, len(executed_acts) * 18)

    # Auto-adjust column widths
    column_widths = {1: 16, 2: 30, 3: 16, 4: 22, 5: 50, 6: 35}
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    if record_id and records:
        user_suffix = f"_solicitud_{records[0].get('solicitud_num', record_id)}"
    elif username:
        user_suffix = f"_{username}"
    else:
        user_suffix = "_global"

    filename = f"reporte_lnet{user_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.get("/api/records/{record_id}/pdf", response_class=HTMLResponse)
def get_record_pdf(record_id: str):
    records = store.load_records()
    record = next((r for r in records if r["id"] == record_id), None)
    if not record:
        raise HTTPException(status_code=404, detail="Registro no encontrado.")

    executed_items = [act for act in record.get("activities", []) if act.get("checked")]
    activities_rows = ""
    for item in executed_items:
        material_name = item.get("description", item.get("name", ""))
        if item.get("detail"):
            material_name += f" ({item.get('detail')})"
        qty = item.get("unid_mts", "N/A")
        activities_rows += f"""
        <tr>
            <td style="padding: 10px; border-bottom: 1px solid #e0e0e0;">{material_name}</td>
            <td style="padding: 10px; border-bottom: 1px solid #e0e0e0; text-align: center; color: #2e7d32; font-weight: bold;">✔ Ejecutado</td>
            <td style="padding: 10px; border-bottom: 1px solid #e0e0e0; text-align: center;">{qty}</td>
        </tr>
        """

    html = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="utf-8">
        <title>Reporte LNet - Solicitud #{record.get('solicitud_num')}</title>
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 30px; color: #333; line-height: 1.5; }}
            .header {{ background: linear-gradient(90deg, #01579b 0%, #0288d1 100%); color: white; padding: 24px; border-radius: 8px; text-align: center; }}
            .card {{ background: #f4f6f9; padding: 18px; border-radius: 8px; margin: 20px 0; border-left: 5px solid #0288d1; font-size: 15px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; background: white; }}
            th {{ background: #0288d1; color: white; padding: 12px; text-align: left; font-size: 14px; }}
            .footer {{ text-align: center; margin-top: 40px; font-size: 12px; color: #777; border-top: 1px solid #eee; padding-top: 15px; }}
            @media print {{
                .no-print {{ display: none !important; }}
                body {{ margin: 0; }}
            }}
        </style>
    </head>
    <body>
        <div class="no-print" style="margin-bottom: 20px; display: flex; justify-content: space-between; align-items: center; background: #e3f2fd; padding: 12px 20px; border-radius: 8px;">
            <strong style="color: #01579b;">Vista Previa de Impresión / Guardado PDF</strong>
            <button onclick="window.print()" style="background: #0288d1; color: white; border: none; padding: 10px 22px; font-size: 15px; font-weight: bold; border-radius: 6px; cursor: pointer; display: flex; align-items: center; gap: 8px;">
                🖨️ Imprimir / Guardar como PDF
            </button>
        </div>

        <div class="header">
            <h2 style="margin: 0; font-size: 24px;">REPORTE DE EJECUCIÓN DE ACTIVIDADES</h2>
            <p style="margin: 5px 0 0 0; opacity: 0.9;">Sistema de Gestión LNet</p>
        </div>

        <div class="card">
            <p style="margin: 4px 0;"><strong>Nro. Solicitud:</strong> #{record.get('solicitud_num')}</p>
            <p style="margin: 4px 0;"><strong>Cliente / Razón Social:</strong> {record.get('client_name')}</p>
            <p style="margin: 4px 0;"><strong>Registrado Por:</strong> {record.get('created_by')} | <strong>Fecha:</strong> {record.get('created_at')}</p>
        </div>

        <h3 style="color: #01579b; border-bottom: 2px solid #0288d1; padding-bottom: 6px;">Materiales y Actividades Ejecutadas</h3>
        <table>
            <thead>
                <tr>
                    <th>CODIGO SAP / MATERIAL UTILIZADO</th>
                    <th style="text-align: center; width: 140px;">ESTADO</th>
                    <th style="text-align: center; width: 140px;">UNID / MTS</th>
                </tr>
            </thead>
            <tbody>
                {activities_rows if activities_rows else '<tr><td colspan="3" style="padding: 15px; text-align: center; color: #666;">No se marcaron actividades ni materiales en esta planilla.</td></tr>'}
            </tbody>
        </table>

        <h3 style="color: #01579b; border-bottom: 2px solid #0288d1; padding-bottom: 6px; margin-top: 30px;">Detalles y Observaciones</h3>
        <div style="background: #fafafa; border: 1px solid #e0e0e0; padding: 15px; border-radius: 6px; min-height: 60px;">
            {record.get('observations') or 'Sin observaciones registradas.'}
        </div>

        <div class="footer">
            Documento de control generado por el Sistema LNet — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        </div>
        <script>
            window.onload = function() {{
                if (window.location.search.includes('print=true')) {{
                    window.print();
                }}
            }}
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html)

@app.post("/api/records")
def create_record(req: RecordCreateRequest):
    # Validate numeric solicitud_num
    if not req.solicitud_num.isdigit():
        raise HTTPException(status_code=400, detail="El Nro. de Solicitud solo debe contener caracteres numéricos.")

    if not req.client_name.strip():
        raise HTTPException(status_code=400, detail="El Nombre / Razón Social es obligatorio.")

    records = store.load_records()
    record_id = str(uuid.uuid4())[:8].upper()
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    email_status = "No enviado"
    email_error = None

    record_data = {
        "id": record_id,
        "solicitud_num": req.solicitud_num.strip(),
        "client_name": req.client_name.strip(),
        "activities": [a.dict() for a in req.activities],
        "observations": req.observations,
        "created_by": req.created_by,
        "created_at": now_str,
        "email_status": email_status
    }

    if req.send_email:
        settings = store.load_settings()
        sender = settings.get("gmail_user")
        app_pw = settings.get("gmail_app_password")

        recipients = req.recipient_email if req.recipient_email else settings.get("default_recipients")

        if not sender or not app_pw:
            record_data["email_status"] = "Fallido (Sin credenciales Gmail configuradas)"
        elif not recipients:
            record_data["email_status"] = "Fallido (Sin destinatario de correo)"
        else:
            try:
                subject = f"[LNet] Ejecución Actividades Solicitud #{req.solicitud_num} - {req.client_name}"
                mailer.send_gmail_email(
                    sender_email=sender,
                    app_password=app_pw,
                    recipient_emails=recipients,
                    subject=subject,
                    record_data=record_data
                )
                record_data["email_status"] = f"Enviado a {recipients}"
            except Exception as e:
                record_data["email_status"] = f"Fallido ({str(e)})"
                email_error = str(e)

    records.insert(0, record_data)
    store.save_records(records)

    res_message = "Registro guardado exitosamente."
    if req.send_email:
        if "Enviado" in record_data["email_status"]:
            res_message += " Correo enviado por Gmail con éxito."
        else:
            res_message += f" Advertencia sobre el correo: {record_data['email_status']}"

    return {
        "message": res_message,
        "record": record_data
    }

@app.post("/api/records/{record_id}/resend")
def resend_record_email(record_id: str, payload: dict = Body(...)):
    records = store.load_records()
    record = next((r for r in records if r["id"] == record_id), None)
    if not record:
        raise HTTPException(status_code=404, detail="Registro no encontrado.")

    settings = store.load_settings()
    sender = settings.get("gmail_user")
    app_pw = settings.get("gmail_app_password")

    recipient = payload.get("recipient_email") or settings.get("default_recipients")
    if not sender or not app_pw:
        raise HTTPException(status_code=400, detail="Faltan credenciales de Gmail en el panel admin.")

    if not recipient:
        raise HTTPException(status_code=400, detail="Debe indicar un destinatario de correo.")

    try:
        subject = f"[LNet] (Reenviado) Ejecución Actividades Solicitud #{record['solicitud_num']} - {record['client_name']}"
        mailer.send_gmail_email(
            sender_email=sender,
            app_password=app_pw,
            recipient_emails=recipient,
            subject=subject,
            record_data=record
        )
        record["email_status"] = f"Enviado a {recipient} ({datetime.now().strftime('%H:%M:%S')})"
        store.save_records(records)
        return {"message": f"Correo reenviado exitosamente a {recipient}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar correo por Gmail: {str(e)}")

# Mount static files from React dist if available
import os
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse

DIST_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend", "dist")
if os.path.exists(DIST_DIR):
    app.mount("/assets", StaticFiles(directory=os.path.join(DIST_DIR, "assets")), name="assets")

    @app.get("/{full_path:path}")
    async def serve_react_app(full_path: str):
        if full_path.startswith("api/"):
            raise HTTPException(status_code=404, detail="API route not found")
        index_file = os.path.join(DIST_DIR, "index.html")
        if os.path.exists(index_file):
            return FileResponse(index_file)
        raise HTTPException(status_code=404, detail="Index file not found")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)

